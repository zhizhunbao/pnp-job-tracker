"""IRCC 开放数据抓取(E12-07 省难度指数,2026-07-20 Frank 拍板「先做 stats 省页卡」):
学签存量(Dec 31 在学口径,非新发)/ 工签存量(TFWP+IMP 分列)/ PNP 登陆数(按省,最新完整年)。
源=open.canada.ca IRCC 官方 XLSX(月更包);数字口径=IRCC 四舍五入到 5、小值 '--' 抑制 → 当 0,
比值用途足够,绝对数不作精算(脚本与前端口径注一致)。配额不在此抓:raw/ircc/pnp_allocations.json 人工核对维护表。
"""
import io
import json
import os
import sys
import urllib.request
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path

import openpyxl

if os.name == "nt":  # 本机控制台 cp1252 打不了 ·/中文;容器由 auto_update 设 PYTHONIOENCODING
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))  # 分域后上一级才是 etl/
import paths

OUT_TR = paths.IRCC / "temp_residents.json"
OUT_PNP = paths.IRCC / "pnp_admissions.json"
print(f"OUT_TR={OUT_TR}\nOUT_PNP={OUT_PNP}", flush=True)

BASE = "https://www.ircc.canada.ca/opendata-donneesouvertes/data/"
SRC = {
    "study": BASE + "EN_ODP_annual-TR-Study-IS_PT_study_level_year_end.xlsx",
    "tfwp": BASE + "EN_ODP_annual-TR-work-TFW_PT_program_year_end.xlsx",
    "imp": BASE + "EN_ODP_annual-TR-work-IMP_PT_program_year_end.xlsx",
    "pr": BASE + "EN_ODP-PR-ProvImmCat.xlsx",
    # 2026-08-03 补:上面 study/tfwp/imp 全是**年末存量**(Dec 31),官方那张表最后一列就停在 2024,
    # 2025 年末存量至今未发(数据集本身 2026-07-21 还更新过,不是僵尸文件)。
    # 但同一数据集的**新发流量**表是**月度**粒度、年份列一直排到 2026 —— 我们一列都没碰过。
    # Frank「是他们没公布还是我们没抓到」→ 存量是他们没发,流量是我们没抓。这一步补后者。
    "study_flow": BASE + "EN_ODP-TR-Study-IS_PT_study_level_sign.xlsx",
}
OUT_FLOW = paths.IRCC / "study_flow.json"
PROV = {
    "Newfoundland and Labrador": "NL", "Prince Edward Island": "PE", "Nova Scotia": "NS",
    "New Brunswick": "NB", "Quebec": "QC", "Ontario": "ON", "Manitoba": "MB",
    "Saskatchewan": "SK", "Alberta": "AB", "British Columbia": "BC",
}

def num(v) -> int:
    s = str(v or "").replace(",", "").strip()
    return int(float(s)) if s and s not in ("--", "") else 0   # '--'=小值抑制 → 0(比值用途可接受)

def fetch(url: str):
    req = urllib.request.Request(url, headers={"User-Agent": "offer2pr-difficulty/1.0"})
    with urllib.request.urlopen(req, timeout=120) as r:
        return openpyxl.load_workbook(BytesIO(r.read()), read_only=True).active

def latest_year_totals(ws) -> tuple[str, dict[str, int]]:
    """年末存量表:省 Total 行(名列 0)× 年份列;取最新有数的年份列。"""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    hdr = next(r for r in rows if any(str(c or "").strip() == "2019" for c in r) or sum(str(c or "").strip().isdigit() for c in r) > 5)
    years = [(i, str(c).strip()) for i, c in enumerate(hdr) if str(c or "").strip()[:2] == "20" and str(c or "").strip().isdigit()]
    out: dict[str, int] = {}
    year_used = ""
    for r in rows:
        if not r:
            continue
        name = str(r[0] or "").replace(" - Total", "").replace(" Total", "").strip()
        if name in PROV:
            # 最新列可能是发布年 YTD/空:从右往左找第一个有值的年份列(全省用同一年,以 ON 首次命中为准)
            for i, y in reversed(years):
                if i < len(r) and str(r[i] or "").strip() not in ("", "--") or (year_used and y == year_used):
                    if not year_used:
                        year_used = y
                    if y == year_used:
                        out[PROV[name]] = num(r[i] if i < len(r) else 0)
                        break
    return year_used, out

def all_year_totals(ws) -> dict[str, dict[str, int]]:
    """年末存量表:省 Total 行 × 全部年份列 → {year: {prov: n}}(2026-08-14 竞争卡年份筛选)。
    列有但整列空(发布年占位)不出;单元格 '--'(小值抑制)按 0(与 latest 同口径)。"""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    hdr = next(r for r in rows if any(str(c or "").strip() == "2019" for c in r) or sum(str(c or "").strip().isdigit() for c in r) > 5)
    years = [(i, str(c).strip()) for i, c in enumerate(hdr) if str(c or "").strip()[:2] == "20" and str(c or "").strip().isdigit()]
    out: dict[str, dict[str, int]] = {}
    for r in rows:
        if not r:
            continue
        name = str(r[0] or "").replace(" - Total", "").replace(" Total", "").strip()
        if name in PROV:
            for i, y in years:
                if i < len(r) and str(r[i] or "").strip() not in ("",):
                    out.setdefault(y, {})[PROV[name]] = num(r[i])
    return {y: m for y, m in out.items() if any(m.values())}


def pnp_latest_full_year(ws) -> tuple[str, dict[str, int]]:
    """PR 按省×类别表:块=类别行…「省 - Total」收尾;取「YYYY Total」最新完整年列的 Provincial Nominee 组行。"""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    # 年总列:hdr 行含「YYYY Total」;倒数第二个=最新完整年(最后一个是进行年 YTD)
    hdr = next(r for r in rows if any("Total" in str(c or "") and str(c or "").strip()[:4].isdigit() for c in r))
    totals = [(i, str(c).strip()[:4]) for i, c in enumerate(hdr) if "Total" in str(c or "") and str(c or "").strip()[:4].isdigit()]
    col, year = totals[-2] if len(totals) >= 2 else totals[-1]
    out: dict[str, int] = {}
    pend: int | None = None
    for r in rows:
        if not r:
            continue
        if any("Provincial Nominee" in str(c or "") for c in r[:4]):
            pend = num(r[col] if col < len(r) else 0)   # 同块成对出现,值相同,留最后一次
        name = str(r[0] or "").replace(" - Total", "").strip()
        if name in PROV and pend is not None:
            out[PROV[name]] = pend
            pend = None
    return year, out

# 表头三层:第 3 行=年(每年起始列,步长 17)、第 4 行=Q1..Q4 与「YYYY Total」、第 5 行=月份。
# 年总计 = 起始列 + 16;进行年(2026)没有年总计列 → 按已有月份列求和作 YTD,并记最后一个有数月份。
MONTHS = ("Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec")


def study_flow(ws) -> dict:
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    yr_row = next(r for r in rows if sum(1 for c in r if str(c or "").strip().isdigit() and str(c or "").strip()[:2] == "20") > 3)
    mo_row = rows[rows.index(yr_row) + 2]
    starts = {str(c).strip(): i for i, c in enumerate(yr_row) if str(c or "").strip().isdigit()}
    def at(r, i):
        return num(r[i]) if i < len(r) else 0
    out: dict[str, dict] = {}
    for r in rows:
        if not r:                                  # 前几行是标题/空行,长度可能为 0
            continue
        name = str(r[0] or "").replace(" Total", "").strip()
        if name not in PROV or "Total" not in str(r[0] or ""):
            continue
        prov: dict[str, dict] = {}
        for y, s0 in starts.items():
            months = [(str(mo_row[k] or "").strip(), k) for k in range(s0, min(s0 + 16, len(mo_row)))]
            months = [(m, k) for m, k in months if m in MONTHS]
            got = [(m, at(r, k)) for m, k in months if k < len(r) and str(r[k] or "").strip() not in ("", "None")]
            if not got:
                continue
            total_col = s0 + 16
            prov[y] = {
                # 年内 12 个月齐 → 用官方年总计列;否则 YTD=已有月份求和,并标出最后一个有数月份
                "n": at(r, total_col) if (len(got) == 12 and total_col < len(r) and at(r, total_col)) else sum(v for _, v in got),
                "complete": len(got) == 12,
                "throughMonth": got[-1][0],
            }
        if prov:
            out[PROV[name]] = prov
    return out


def main() -> None:
    paths.IRCC.mkdir(parents=True, exist_ok=True)
    # fetched = 本轮抓取日(B3-3):要拿来下结论的数据必须知道是哪天的;新鲜度哨兵(check_freshness)也读它
    fetched = datetime.now(timezone.utc).date().isoformat()
    # 年份哨兵(2026-08-14):存量最新年份一变(IRCC 补发 2025)就大声喊 —— 竞争卡年份列会自动亮,
    # 但当年名额(pnp_allocations 人工表)与竞争比口径要人跟着核,静默自愈=没人知道该去补
    old_years: dict[str, str] = {}
    if OUT_TR.exists():
        try:
            _old = json.loads(OUT_TR.read_text(encoding="utf-8"))
            old_years = {k: str((_old.get(k) or {}).get("year") or "") for k in ("study", "tfwp", "imp")}
        except Exception:  # noqa: BLE001  旧表坏了不影响本轮抓取
            pass
    tr: dict = {"fetched": fetched, "source": {k: SRC[k] for k in ("study", "tfwp", "imp")}, "note": "IRCC 年末存量(Dec 31 holders);数值官方四舍五入到 5,'--' 小值抑制当 0"}
    for key in ("study", "tfwp", "imp"):
        ws = fetch(SRC[key])
        year, totals = latest_year_totals(ws)
        if old_years.get(key) and old_years[key] != year:
            print(f"! {key} 年末存量最新年份 {old_years[key]} → {year}:IRCC 补发了新年份 —— "
                  f"去 pnp_allocations.json 核对 {year} 名额,并复核竞争卡该年列", flush=True)
        # byYear:全年份序列(2026-08-14 竞争卡年份筛选)。同一张表只抓一次 —— read_only 工作簿
        # iter_rows 可重复遍历,两个解析器各自扫行
        by_year = all_year_totals(ws)
        tr[key] = {"year": year, "byProv": totals, "byYear": by_year}
        print(f"{key}: {year} · {len(totals)} 省 · ON={totals.get('ON')} · 序列 {min(by_year)}–{max(by_year)}", flush=True)
    OUT_TR.write_text(json.dumps(tr, ensure_ascii=False, indent=1), encoding="utf-8")

    year, pnp = pnp_latest_full_year(fetch(SRC["pr"]))
    OUT_PNP.write_text(json.dumps({"fetched": fetched, "source": SRC["pr"], "year": year, "byProv": pnp,
                                   "note": "PNP 类别 PR 登陆数(含随行家属,人头口径)最新完整年"}, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"pnp admissions: {year} · {len(pnp)} 省 · ON={pnp.get('ON')}", flush=True)

    flow = study_flow(fetch(SRC["study_flow"]))
    OUT_FLOW.write_text(json.dumps({
        "fetched": fetched, "source": SRC["study_flow"], "byProv": flow,
        "note": ("新发学签**流量**(按许可生效月份,非年末存量)。月度粒度,进行年为 YTD(complete=false 时 "
                 "n 是已公布月份求和,throughMonth 是最后一个有数月份)。与存量口径不可混用:"
                 "存量=在库人数(竞争比分母),流量=当期新增趋势。"),
    }, ensure_ascii=False, indent=1), encoding="utf-8")
    yrs = sorted({y for p in flow.values() for y in p})
    on = flow.get("ON", {})
    tail = "  ".join(
        f"{y}={on[y]['n']:,}" + ("" if on[y]["complete"] else " (至 " + on[y]["throughMonth"] + ")")
        for y in yrs[-3:] if y in on)
    print(f"study flow: {len(flow)} 省 · 年份 {yrs[0]}–{yrs[-1]} · ON {tail}", flush=True)

if __name__ == "__main__":
    main()
