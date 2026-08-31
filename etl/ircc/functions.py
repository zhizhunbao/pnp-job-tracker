"""
ircc 域函数 —— 全部行为住这(照 company/pnp 全溶样张,方言律全集见
docs/design/etl分域-20260829.md §4)。

原 5 个步骤文件 + 一个横切步 2026-08-30 批C 溶入本文件,一步一段(段横幅三行框 + N. 编号,
与 constants.py 同名同序镜像),各段入口函数与原脚本同名、一律零参,门(main.py)直调。
⚠ 段5 的 clean/04e_difficulty.py **不溶**:它是清洗横切层文件(一个关注点一个脚本、跨源生效),
本域只写一个零参包装函数把它按序跑起来(subprocess 先例:load.functions 的 pg_dump、
wages.functions 的 curl)。
**零字符串令**:字面量全住 constants(文案 *_TPL 模板、JSON 键 K_ 词族、官方原句在 *_RULES);
**显式循环令**:禁推导/genexp/lambda;**内嵌禁令**:内部函数出户成顶层具名函数;
**一参令**:函数至多一参,多入参收 scheme 的 XxxIn dataclass,多返回值收 XxxOut。
日志口径:域内不裸 print,报数走 log.functions.say;各步原有的「✗ …(保留旧表)」「! …」行
原样保留(auto_update 按 ✗/! 行首升 ERROR 级,这就是本域的错误通道);
原来静默 pass 的 catch 补 err() 留痕(永不吞异常令)。
⚠ 原 scrape_ircc_stats / scrape_statcan_tr_prov 头部的 Windows 控制台 TextIOWrapper 兜底
(`if os.name == "nt": sys.stdout = io.TextIOWrapper(...)`,注「本机控制台 cp1252 打不了
·/中文;容器由 auto_update 设 PYTHONIOENCODING」)随裸 print 一同退役 —— 出口已统一成
log.functions(loguru),与 pnp/company/dli 等全溶域一致;本机跑一律带 PYTHONIOENCODING=utf-8。
硬闸口径不变:自校未过 → 保留旧表 + sys.exit(1),门见 SystemExit 直接中止本轮
(与旧 _steps 跑子进程「一步失败即中止」同语义)。
依赖单边:本文件 → constants/scheme + 基础设施叶(paths / log / fetch / crawl)。
"""
import json
import subprocess
import sys
import urllib.request
from datetime import date, datetime, timezone
from io import BytesIO

import httpx
import openpyxl
from bs4 import BeautifulSoup

import paths
from log.functions import err, say
from fetch.constants import HDR_UA, PARSER_HTML, WS_RE
from crawl.functions import convert_md
from crawl.scheme import ConvertIn
from ircc.constants import (
    BLANK_VALUES, COMMA, COORD_SEP, COORD_TPL, DIFFICULTY_FAIL_TPL,
    DIFFICULTY_SCRIPT, ENC_UTF8, FEE_FACTOR, FEE_OP, FEE_UNIT, FEES_BIO_ITEMS, FEES_BULLET_TPL,
    FEES_DONE_TPL, FEES_DROP_TAGS, FEES_FAIL_HEADER, FEES_ITEMS, FEES_NOTE, FEES_PRINT_OUT_TPL,
    FEES_PROBLEM_ITEM_TPL, FEES_PROBLEM_NO_SECTION_TPL, FEES_PROBLEM_RPRF_TPL, FEES_PROGRAM,
    FEES_SECTION, FEES_SECTION_BIO, FEES_SEG_LEN, FEES_SOURCE, FEES_TIMEOUT_S, FEES_UA, FEES_URL,
    FEES_VALUE_CLIP, FLOW_BLANKS, FLOW_TAIL_PARTIAL_TPL, FLOW_TAIL_SEP, FLOW_TAIL_TPL,
    FLOW_TAIL_YEARS, FLOW_YEAR_ROW_MIN, GEO_DIM, HDR_MIN_DIGITS, HDR_PROBE_YEAR, INDENT_1,
    K_APPLIES_AREA, K_APPLIES_NOC, K_APPLIES_TEER, K_BASIS, K_BY_PROV, K_BY_YEAR, K_COMPLETE,
    K_COORDINATE, K_DIMENSION, K_DIMENSION_NAME_EN, K_EFFECTIVE, K_EXCLUDES_NOC, K_FACTOR,
    K_FAMILY_SIZE, K_FETCHED, K_FETCHED_AT, K_GAP_TO_TARGET, K_LABEL, K_LATEST, K_LATEST_N,
    K_LATEST_REF_PER, K_MEMBER, K_MEMBER_ID, K_MEMBER_NAME_EN, K_N, K_NOTE, K_NPR, K_OBJECT, K_OP,
    K_PAGE, K_PEAK, K_PER_QUARTER_CHANGE, K_POPULATION, K_PRODUCT_ID, K_PROGRAM, K_PROVINCE,
    K_QUARTERS, K_QUARTERS_TO_TARGET, K_QUOTE, K_REF_PER, K_REQUIREMENTS, K_SECTION, K_SHARE,
    K_SOURCE, K_STATUS, K_STREAM, K_STUDY_ONLY, K_SUBJECT, K_TARGET, K_THROUGH_MONTH, K_TYPES,
    K_UNIT, K_URL, K_VALUE, K_VALUE_TEXT, K_VECTOR_DATA_POINT, K_VECTOR_ID, K_WORK_ONLY,
    K_WORK_STUDY, K_YEAR, MONTH_SPAN, MONTHS, MONTHS_FULL, NPR_DONE_TPL, NPR_FAIL_TPL,
    NPR_LATEST_TPL, NPR_MIN_QUARTERS, NPR_NOTE, NPR_PEAK_TPL, NPR_PRINT_OUT_TPL, NPR_QUARTERS,
    NPR_SPAN, NPR_SPEED_TPL, NPR_SRC_URLS, NPR_TARGET, NPR_TIMEOUT_S, NPR_TOO_FEW_TPL, NPR_UA,
    NPR_WDS, OUT_FEES, OUT_FLOW, OUT_NPR, OUT_PGWP, OUT_PNP, OUT_TR, OUT_TR_PROV, PCT_SCALE,
    PGWP_DONE_TPL, PGWP_MISSING_ROW_TPL, PGWP_MISSING_TPL, PGWP_NOTE, PGWP_PAGE_ABOUT,
    PGWP_PAGE_ELIG, PGWP_PRINT_OUT_TPL, PGWP_PROGRAM, PGWP_QUOTE_CLIP, PGWP_RULES, PGWP_STAR,
    PGWP_TIMEOUT_S, PGWP_UA, PGWP_URL_ABOUT, PGWP_URL_ELIG, PNP_CATEGORY_WORD, PROV_CODE, PROV_ON,
    PROVINCE_FED, QUARTERS_ROUND, QUOTE_CURLY_LEFT, QUOTE_CURLY_RIGHT, QUOTE_STRAIGHT, SHARE_ROUND,
    SPACE, STATS_FLOW_NOTE, STATS_FLOW_TPL, STATS_NO_FLOW_HEADER, STATS_NO_HEADER,
    STATS_NO_PNP_HEADER, STATS_PNP_NOTE, STATS_PNP_TPL, STATS_PRINT_OUT_TPL, STATS_SRC,
    STATS_STOCK_TPL, STATS_TIMEOUT_S, STATS_TR_NOTE, STATS_UA, STATS_YEAR_ALERT_TPL,
    SRC_PR, SRC_STUDY_FLOW, STATUS_SUCCESS, STOCK_KEYS, STREAM_PRINCIPAL,
    STREAM_PRINCIPAL_NO_RPRF, SUBJECT_APPLICANT, TIMESPEC_SECONDS, TOTAL_DASH_SUFFIX, TOTAL_SUFFIX,
    TOTAL_WORD, TRP_COORD_FAIL_TPL, TRP_DATA_TIMEOUT_S, TRP_DATA_URL, TRP_DIM_FAIL_TPL,
    TRP_DONE_TPL, TRP_FAIL_TPL, TRP_META_TIMEOUT_S, TRP_META_URL, TRP_MIN_PROV, TRP_MIN_TYPES,
    TRP_NOTE, TRP_ON_MIN, TRP_PID, TRP_PRINT_OUT_TPL, TRP_QUARTERS, TRP_ROW_TPL, TRP_SANITY_FAIL,
    TRP_SRC_URL, TRP_TYPES, TRP_UA, TYPE_DIM_WORD, V_NPR, V_POP, WDS_STATUS_FAIL_TPL,
    WDS_VECTOR_FAIL_TPL, YEAR_PREFIX, YEAR_TOTAL_OFFSET,
)
from ircc.scheme import (
    ByProvIn, CellAtIn, CoordIn, FailIn, FeeRowIn, FlowGotIn, FlowMonthsIn, FlowTailIn,
    FlowYearIn, ItemsOut, MemberIds, NprRowsIn, PgwpReqIn, QuartersIn, SectionItemsIn, YearTotals,
)

# =========================================================================
# 1. 共享词汇(≥2 段消费:落盘日戳 / 取格 / 自校失败出口)
# =========================================================================


def today_iso() -> str:
    """今天(本地日期,ISO)—— NPR / 分省存量 / PGWP / 规费四段 fetched 的口径。"""
    return date.today().isoformat()


def cell_text(v: object) -> str:
    """一格 → 去空白的文本(缺格当空串)——「str(c or '').strip()」的唯一出口。"""
    return str(v or "").strip()


def fail_keep_old(x: FailIn) -> None:
    """自校未过:抬头 + 逐条明细 → 退出码 1(**保留旧表不覆盖**)。

    门直调本域函数,SystemExit 不被 `except Exception` 接住 —— 与旧 _steps 跑子进程时
    「某步 exit 1 即中止本轮」逐字同语义。
    """
    say(x.header)
    for line in x.lines:
        say(line)
    sys.exit(1)


# =========================================================================
# 2. IRCC 开放数据:学签/工签年末存量 + PNP 登陆数 + 新发学签流量
# =========================================================================


def utc_today_iso() -> str:
    """本轮抓取日(UTC 日期,ISO)—— 段2 三张表 fetched 的口径(与其余段的本地日不同,原值照搬)。"""
    return datetime.now(timezone.utc).date().isoformat()


def num(v: object) -> int:
    """一格 → 整数;空格与小值抑制记号 '--' 一律当 0(比值用途可接受)。"""
    s = str(v or "").replace(COMMA, "").strip()
    if s in BLANK_VALUES:
        return 0
    return int(float(s))


def cell_at(x: CellAtIn) -> int:
    """行内按列下标取整数(越界当 0)—— 原 study_flow 的内嵌 at() 出户。"""
    if x.index < len(x.row):
        return num(x.row[x.index])
    return 0


def sheet_rows(ws: object) -> list:
    """工作表 → 行清单(read_only 工作簿 iter_rows 可重复遍历,两个解析器各自扫行)。"""
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append(list(r))
    return rows


def is_year_header_row(r: list) -> bool:
    """年份表头行:含探针年份,或纯数字格超过防线。"""
    digits = 0
    for c in r:
        t = cell_text(c)
        if t == HDR_PROBE_YEAR:
            return True
        if t.isdigit():
            digits += 1
    return digits > HDR_MIN_DIGITS


def year_header_of(rows: list) -> list:
    """找出年份表头行;找不到即报错(原 next() 的 StopIteration,批C 换成带话的报错)。"""
    for r in rows:
        if is_year_header_row(r):
            return r
    raise RuntimeError(STATS_NO_HEADER)


def year_columns_of(hdr: list) -> list:
    """表头 → [(列下标, 年份)] —— 只认 20xx 的纯数字列。"""
    out = []
    i = 0
    for c in hdr:
        t = cell_text(c)
        if t[:2] == YEAR_PREFIX and t.isdigit():
            out.append((i, t))
        i += 1
    return out


def prov_name_of(cell: object) -> str:
    """省行名:剥掉「 - Total」/「 Total」尾巴。"""
    return str(cell or "").replace(TOTAL_DASH_SUFFIX, "").replace(TOTAL_SUFFIX, "").strip()


def latest_year_totals(ws: object) -> YearTotals:
    """年末存量表:省 Total 行(名列 0)× 年份列;取最新有数的年份列。"""
    rows = sheet_rows(ws)
    years = year_columns_of(year_header_of(rows))
    out: dict = {}
    year_used = ""
    for r in rows:
        if len(r) == 0:
            continue
        name = prov_name_of(r[0])
        if name not in PROV_CODE:
            continue
        for i, y in reversed(years):
            has_value = i < len(r) and cell_text(r[i]) not in BLANK_VALUES
            if has_value or (year_used != "" and y == year_used):
                if year_used == "":
                    year_used = y
                if y == year_used:
                    out[PROV_CODE[name]] = cell_at(CellAtIn(row=r, index=i))
                    break
    return YearTotals(year=year_used, by_prov=out)


def all_year_totals(ws: object) -> dict:
    """年末存量表:省 Total 行 × 全部年份列 → {year: {prov: n}}(2026-08-14 竞争卡年份筛选)。

    列有但整列空(发布年占位)不出;单元格 '--'(小值抑制)按 0(与 latest 同口径)。
    """
    rows = sheet_rows(ws)
    years = year_columns_of(year_header_of(rows))
    out: dict = {}
    for r in rows:
        if len(r) == 0:
            continue
        name = prov_name_of(r[0])
        if name not in PROV_CODE:
            continue
        for i, y in years:
            if i < len(r) and cell_text(r[i]) != "":
                out.setdefault(y, {})[PROV_CODE[name]] = num(r[i])
    kept: dict = {}
    for y, m in out.items():
        if any(m.values()):
            kept[y] = m
    return kept


def is_year_total_row(r: list) -> bool:
    """PR 表的年总列表头行:某格含「Total」且前四字是数字。"""
    for c in r:
        t = str(c or "")
        if TOTAL_WORD in t and t.strip()[:4].isdigit():
            return True
    return False


def year_total_columns_of(hdr: list) -> list:
    """表头 → [(列下标, 年份)]:含「YYYY Total」的列。"""
    out = []
    i = 0
    for c in hdr:
        t = str(c or "")
        if TOTAL_WORD in t and t.strip()[:4].isdigit():
            out.append((i, t.strip()[:4]))
        i += 1
    return out


def has_pnp_cell(r: list) -> bool:
    """本行前四格里有没有「Provincial Nominee」类别名。"""
    for c in r[:4]:
        if PNP_CATEGORY_WORD in str(c or ""):
            return True
    return False


def pnp_latest_full_year(ws: object) -> YearTotals:
    """PR 按省×类别表:块=类别行…「省 - Total」收尾;取「YYYY Total」最新完整年列的
    Provincial Nominee 组行。

    年总列:hdr 行含「YYYY Total」;倒数第二个=最新完整年(最后一个是进行年 YTD)。
    """
    rows = sheet_rows(ws)
    hdr = None
    for r in rows:
        if is_year_total_row(r):
            hdr = r
            break
    if hdr is None:
        raise RuntimeError(STATS_NO_PNP_HEADER)
    totals = year_total_columns_of(hdr)
    col, year = totals[-1]
    if len(totals) >= 2:
        col, year = totals[-2]
    out: dict = {}
    pend = None
    for r in rows:
        if len(r) == 0:
            continue
        if has_pnp_cell(r):
            pend = cell_at(CellAtIn(row=r, index=col))
        name = str(r[0] or "").replace(TOTAL_DASH_SUFFIX, "").strip()
        if name in PROV_CODE and pend is not None:
            out[PROV_CODE[name]] = pend
            pend = None
    return YearTotals(year=year, by_prov=out)


def is_flow_year_row(r: list) -> bool:
    """流量表的年份行:20xx 纯数字格超过防线。"""
    n = 0
    for c in r:
        t = cell_text(c)
        if t.isdigit() and t[:2] == YEAR_PREFIX:
            n += 1
    return n > FLOW_YEAR_ROW_MIN


def year_starts_of(yr_row: list) -> dict:
    """年份行 → {年: 起始列}。"""
    out: dict = {}
    i = 0
    for c in yr_row:
        if cell_text(c).isdigit():
            out[str(c).strip()] = i
        i += 1
    return out


def flow_months_of(x: FlowMonthsIn) -> list:
    """本年度的月份列 [(月名, 列下标)] —— 只认 MONTHS 里的名字。"""
    out = []
    end = min(x.start + MONTH_SPAN, len(x.mo_row))
    for k in range(x.start, end):
        m = cell_text(x.mo_row[k])
        if m in MONTHS:
            out.append((m, k))
    return out


def flow_got_of(x: FlowGotIn) -> list:
    """本年度有数的月份 [(月名, 值)]。"""
    out = []
    for m, k in x.months:
        if k < len(x.row) and cell_text(x.row[k]) not in FLOW_BLANKS:
            out.append((m, cell_at(CellAtIn(row=x.row, index=k))))
    return out


def flow_year_row(x: FlowYearIn) -> dict:
    """本年度一块:年内 12 个月齐 → 用官方年总计列;否则 YTD=已有月份求和,并标出最后一个有数月份。"""
    total_col = x.start + YEAR_TOTAL_OFFSET
    total = 0
    for _m, v in x.got:
        total += v
    if len(x.got) == MONTHS_FULL and total_col < len(x.row):
        official = cell_at(CellAtIn(row=x.row, index=total_col))
        if official:
            total = official
    return {K_N: total, K_COMPLETE: len(x.got) == MONTHS_FULL, K_THROUGH_MONTH: x.got[-1][0]}


def study_flow(ws: object) -> dict:
    """新发学签流量表 → {省码: {年: 年块}}(省 Total 行 × 年 × 月)。"""
    rows = sheet_rows(ws)
    yr_row = None
    for r in rows:
        if is_flow_year_row(r):
            yr_row = r
            break
    if yr_row is None:
        raise RuntimeError(STATS_NO_FLOW_HEADER)
    mo_row = rows[rows.index(yr_row) + 2]
    starts = year_starts_of(yr_row)
    out: dict = {}
    for r in rows:
        if len(r) == 0:
            continue
        raw = str(r[0] or "")
        name = raw.replace(TOTAL_SUFFIX, "").strip()
        if name not in PROV_CODE or TOTAL_WORD not in raw:
            continue
        prov: dict = {}
        for y, s0 in starts.items():
            got = flow_got_of(FlowGotIn(row=r, months=flow_months_of(
                FlowMonthsIn(mo_row=mo_row, start=s0))))
            if len(got) == 0:
                continue
            prov[y] = flow_year_row(FlowYearIn(row=r, start=s0, got=got))
        if len(prov) > 0:
            out[PROV_CODE[name]] = prov
    return out


def flow_years_of(flow: dict) -> list:
    """全部年份(升序)。"""
    ys = set()
    for p in flow.values():
        for y in p:
            ys.add(y)
    return sorted(ys)


def flow_tail_of(x: FlowTailIn) -> str:
    """收尾报数的 ON 尾巴:最近几年,进行年标到哪个月。"""
    on = x.flow.get(PROV_ON)
    if on is None:
        on = {}
    parts = []
    for y in x.years[-FLOW_TAIL_YEARS:]:
        if y not in on:
            continue
        piece = FLOW_TAIL_TPL.format(year=y, n=on[y][K_N])
        if on[y][K_COMPLETE] is False:
            piece += FLOW_TAIL_PARTIAL_TPL.format(month=on[y][K_THROUGH_MONTH])
        parts.append(piece)
    return FLOW_TAIL_SEP.join(parts)


def fetch_sheet(url: str) -> object:
    """下载一张官方 XLSX → 活动工作表(read_only,省内存)。"""
    req = urllib.request.Request(url, headers={HDR_UA: STATS_UA})
    with urllib.request.urlopen(req, timeout=STATS_TIMEOUT_S) as r:
        return openpyxl.load_workbook(BytesIO(r.read()), read_only=True).active


def old_stock_years() -> dict:
    """旧表里三种存量的最新年份 —— 年份哨兵的比对基准(旧表坏了不影响本轮抓取)。"""
    out: dict = {}
    if not OUT_TR.exists():
        return out
    try:
        old = json.loads(OUT_TR.read_text(encoding=ENC_UTF8))
    except Exception as e:  # noqa: BLE001
        err(OUT_TR, e)
        return out
    for k in STOCK_KEYS:
        block = old.get(k)
        if block is None:
            block = {}
        out[k] = str(block.get(K_YEAR) or "")
    return out


def write_stock_table(fetched: str) -> None:
    """三张年末存量表 → OUT_TR(含年份哨兵告警)。"""
    old_years = old_stock_years()
    source: dict = {}
    for k in STOCK_KEYS:
        source[k] = STATS_SRC[k]
    tr: dict = {K_FETCHED: fetched, K_SOURCE: source, K_NOTE: STATS_TR_NOTE}
    for key in STOCK_KEYS:
        ws = fetch_sheet(STATS_SRC[key])
        got = latest_year_totals(ws)
        old = old_years.get(key, "")
        if old != "" and old != got.year:
            say(STATS_YEAR_ALERT_TPL.format(key=key, old=old, year=got.year))
        by_year = all_year_totals(ws)
        tr[key] = {K_YEAR: got.year, K_BY_PROV: got.by_prov, K_BY_YEAR: by_year}
        say(STATS_STOCK_TPL.format(key=key, year=got.year, n=len(got.by_prov),
                                   on=got.by_prov.get(PROV_ON), first=min(by_year),
                                   last=max(by_year)))
    paths.write_json(paths.WriteJsonIn(path=OUT_TR, payload=tr, indent=INDENT_1))


def write_pnp_admissions(fetched: str) -> None:
    """PR 按省×类别表 → OUT_PNP(最新完整年的 PNP 组行)。"""
    pnp = pnp_latest_full_year(fetch_sheet(STATS_SRC[SRC_PR]))
    paths.write_json(paths.WriteJsonIn(path=OUT_PNP, payload={
        K_FETCHED: fetched, K_SOURCE: STATS_SRC[SRC_PR], K_YEAR: pnp.year,
        K_BY_PROV: pnp.by_prov, K_NOTE: STATS_PNP_NOTE,
    }, indent=INDENT_1))
    say(STATS_PNP_TPL.format(year=pnp.year, n=len(pnp.by_prov), on=pnp.by_prov.get(PROV_ON)))


def write_study_flow(fetched: str) -> None:
    """新发学签流量表 → OUT_FLOW(月度粒度,进行年为 YTD)。"""
    flow = study_flow(fetch_sheet(STATS_SRC[SRC_STUDY_FLOW]))
    paths.write_json(paths.WriteJsonIn(path=OUT_FLOW, payload={
        K_FETCHED: fetched, K_SOURCE: STATS_SRC[SRC_STUDY_FLOW], K_BY_PROV: flow,
        K_NOTE: STATS_FLOW_NOTE,
    }, indent=INDENT_1))
    years = flow_years_of(flow)
    say(STATS_FLOW_TPL.format(n=len(flow), first=years[0], last=years[-1],
                              tail=flow_tail_of(FlowTailIn(flow=flow, years=years))))


def scrape_ircc_stats() -> None:
    """IRCC 开放数据抓取(E12-07 省难度指数,2026-07-20 Frank 拍板「先做 stats 省页卡」)。

    学签存量(Dec 31 在学口径,非新发)/ 工签存量(TFWP+IMP 分列)/ PNP 登陆数(按省,
    最新完整年)/ 新发学签流量。源=open.canada.ca IRCC 官方 XLSX(月更包);
    数字口径=IRCC 四舍五入到 5、小值 '--' 抑制 → 当 0,比值用途足够,绝对数不作精算
    (脚本与前端口径注一致)。
    配额不在此抓:raw/ircc/pnp_allocations.json 是人工核对维护表。
    """
    say(STATS_PRINT_OUT_TPL.format(tr=OUT_TR, pnp=OUT_PNP))
    paths.IRCC.mkdir(parents=True, exist_ok=True)
    fetched = utc_today_iso()
    write_stock_table(fetched)
    write_pnp_admissions(fetched)
    write_study_flow(fetched)


# =========================================================================
# 3. NPR 占总人口比(联邦「临时人口降到 5%」目标的唯一可核验刻度)
# =========================================================================


def series(vector: int) -> dict:
    """一条 WDS 向量 → {季度参考日: 值}(非 SUCCESS 即抛,交调用方保留旧表)。"""
    r = httpx.post(NPR_WDS, json=[{K_VECTOR_ID: vector, K_LATEST_N: NPR_QUARTERS}],
                   headers={HDR_UA: NPR_UA}, timeout=NPR_TIMEOUT_S)
    r.raise_for_status()
    blk = r.json()[0]
    if blk.get(K_STATUS) != STATUS_SUCCESS:
        raise RuntimeError(WDS_VECTOR_FAIL_TPL.format(status=blk.get(K_STATUS), vector=vector))
    out: dict = {}
    for p in blk[K_OBJECT][K_VECTOR_DATA_POINT]:
        if p.get(K_VALUE) is not None:
            out[p[K_REF_PER]] = float(p[K_VALUE])
    return out


def npr_rows_of(x: NprRowsIn) -> list:
    """两条序列 → 季度行(只留两边都有的季度)。"""
    rows = []
    for q in sorted(x.npr):
        if q not in x.pop:
            continue
        rows.append({K_REF_PER: q, K_POPULATION: int(x.pop[q]), K_NPR: int(x.npr[q]),
                     K_SHARE: round(x.npr[q] / x.pop[q], SHARE_ROUND)})
    return rows


def share_of(row: dict) -> float:
    """峰值排序键(原 lambda 出户成具名)。"""
    return row[K_SHARE]


def quarters_to_target_of(x: QuartersIn) -> float | None:
    """按最近四季降速线性外推到 5% 还要几个季度;没在降 → None(不外推)。"""
    if x.per_q < 0:
        return round((NPR_TARGET - x.share) / x.per_q, QUARTERS_ROUND)
    return None


def scrape_statcan_npr() -> None:
    """StatCan 非永久居民(NPR)占总人口比 → raw/ircc/npr_share.json。

    IN : StatCan WDS(免密钥 REST):v1=加拿大季度总人口 / v1566927590=非永久居民(NPR)总数
    OUT: raw/ircc/npr_share.json(季度序列 + 最新占比 + 距 5% 目标的人数缺口)
    """
    say(NPR_PRINT_OUT_TPL.format(path=OUT_NPR))
    paths.IRCC.mkdir(parents=True, exist_ok=True)
    try:
        pop = series(V_POP)
        npr = series(V_NPR)
    except Exception as e:  # noqa: BLE001
        say(NPR_FAIL_TPL.format(name=type(e).__name__, detail=e))
        return
    rows = npr_rows_of(NprRowsIn(pop=pop, npr=npr))
    if len(rows) < NPR_MIN_QUARTERS:
        say(NPR_TOO_FEW_TPL.format(n=len(rows)))
        return
    latest = rows[-1]
    peak = max(rows, key=share_of)
    span = rows
    if len(rows) >= NPR_SPAN:
        span = rows[-NPR_SPAN:]
    per_q = (span[-1][K_SHARE] - span[0][K_SHARE]) / max(len(span) - 1, 1)
    gap_people = int(latest[K_NPR] - latest[K_POPULATION] * NPR_TARGET)
    quarters = quarters_to_target_of(QuartersIn(share=latest[K_SHARE], per_q=per_q))
    paths.write_json(paths.WriteJsonIn(path=OUT_NPR, payload={
        K_SOURCE: NPR_SRC_URLS, K_FETCHED: today_iso(),
        K_FETCHED_AT: datetime.now(timezone.utc).isoformat(timespec=TIMESPEC_SECONDS),
        K_TARGET: NPR_TARGET, K_QUARTERS: rows,
        K_LATEST: latest, K_PEAK: peak,
        K_PER_QUARTER_CHANGE: round(per_q, SHARE_ROUND),
        K_GAP_TO_TARGET: gap_people,
        K_QUARTERS_TO_TARGET: quarters,
        K_NOTE: NPR_NOTE,
    }, indent=INDENT_1))
    say(NPR_DONE_TPL.format(n=len(rows), out=OUT_NPR.name))
    say(NPR_LATEST_TPL.format(ref=latest[K_REF_PER], pct=latest[K_SHARE] * PCT_SCALE,
                              npr=latest[K_NPR], pop=latest[K_POPULATION]))
    say(NPR_PEAK_TPL.format(ref=peak[K_REF_PER], pct=peak[K_SHARE] * PCT_SCALE, gap=gap_people))
    say(NPR_SPEED_TPL.format(per=per_q * PCT_SCALE, quarters=quarters))


# =========================================================================
# 4. StatCan 分省临时居民存量(IRCC 年末存量停在 2024 后唯一的官方分省刻度)
# =========================================================================


def member_ids() -> MemberIds:
    """metadata 解析省/证型的 memberId(不写死:StatCan 重排成员时坐标会静默错位)。"""
    r = httpx.post(TRP_META_URL, json=[{K_PRODUCT_ID: TRP_PID}], headers={HDR_UA: TRP_UA},
                   timeout=TRP_META_TIMEOUT_S)
    r.raise_for_status()
    geo: dict = {}
    typ: dict = {}
    for d in r.json()[0][K_OBJECT][K_DIMENSION]:
        name = d[K_DIMENSION_NAME_EN]
        if name == GEO_DIM:
            for m in d[K_MEMBER]:
                geo[m[K_MEMBER_NAME_EN]] = int(m[K_MEMBER_ID])
        if TYPE_DIM_WORD in name.lower():
            for m in d[K_MEMBER]:
                typ[m[K_MEMBER_NAME_EN]] = int(m[K_MEMBER_ID])
    geo_ids: dict = {}
    for name, code in PROV_CODE.items():
        if name in geo:
            geo_ids[code] = geo[name]
    typ_ids: dict = {}
    for key, name in TRP_TYPES.items():
        if name in typ:
            typ_ids[key] = typ[name]
    return MemberIds(geo=geo_ids, types=typ_ids)


def coord_of(x: CoordIn) -> str:
    """(省 memberId, 证型 memberId) → WDS 十维坐标串。"""
    return COORD_TPL.format(geo=x.geo, typ=x.typ)


def tr_prov_requests(ids: MemberIds) -> list:
    """省 × 证型 的取数请求(一发全取)。"""
    reqs = []
    for g in ids.geo.values():
        for t in ids.types.values():
            reqs.append({K_PRODUCT_ID: TRP_PID, K_COORDINATE: coord_of(CoordIn(geo=g, typ=t)),
                         K_LATEST_N: TRP_QUARTERS})
    return reqs


def tr_prov_by_prov(x: ByProvIn) -> dict:
    """响应块 → {省码: {季度: {证型: 值}}}。

    响应块**不按请求顺序**回来(实测乱序)—— 只能从块自带 coordinate 反解 (省, 证型)。
    """
    prov_of: dict = {}
    for p, g in x.ids.geo.items():
        prov_of[g] = p
    key_of: dict = {}
    for k, t in x.ids.types.items():
        key_of[t] = k
    by_prov: dict = {}
    for blk in x.blocks:
        if blk.get(K_STATUS) != STATUS_SUCCESS:
            raise RuntimeError(WDS_STATUS_FAIL_TPL.format(status=blk.get(K_STATUS)))
        o = blk[K_OBJECT]
        parts = o[K_COORDINATE].split(COORD_SEP)
        prov = prov_of.get(int(parts[0]))
        key = key_of.get(int(parts[1]))
        if prov is None or key is None:
            raise RuntimeError(TRP_COORD_FAIL_TPL.format(coord=o[K_COORDINATE]))
        for p in o[K_VECTOR_DATA_POINT]:
            if p.get(K_VALUE) is not None:
                by_prov.setdefault(prov, {}).setdefault(p[K_REF_PER], {})[key] = int(p[K_VALUE])
    return by_prov


def latest_ref_of(by_prov: dict) -> str:
    """全部省里最新的那个季度参考日。"""
    quarters = []
    for p in by_prov.values():
        for q in p:
            quarters.append(q)
    return max(quarters)


def scrape_statcan_tr_prov() -> None:
    """StatCan 分省临时居民存量(季度)→ raw/ircc/statcan_tr_prov.json。

    IN : StatCan WDS getCubeMetadata + getDataFromCubePidCoordAndLatestNPeriods (pid 17100121)
    OUT: raw/ircc/statcan_tr_prov.json
    抓取失败 / 维度缺位 / 量级失真 → 保留旧表(宁可留旧也不留空)。
    """
    say(TRP_PRINT_OUT_TPL.format(path=OUT_TR_PROV))
    paths.IRCC.mkdir(parents=True, exist_ok=True)
    try:
        ids = member_ids()
        if len(ids.geo) < TRP_MIN_PROV or len(ids.types) < TRP_MIN_TYPES:
            raise RuntimeError(TRP_DIM_FAIL_TPL.format(geo=len(ids.geo), typ=len(ids.types)))
        r = httpx.post(TRP_DATA_URL, json=tr_prov_requests(ids), headers={HDR_UA: TRP_UA},
                       timeout=TRP_DATA_TIMEOUT_S)
        r.raise_for_status()
        by_prov = tr_prov_by_prov(ByProvIn(blocks=r.json(), ids=ids))
        latest = latest_ref_of(by_prov)
        checked = by_prov.get(PROV_ON, {}).get(latest, {})
        if (checked.get(K_STUDY_ONLY) or 0) < TRP_ON_MIN:
            raise RuntimeError(TRP_SANITY_FAIL)
    except Exception as e:  # noqa: BLE001
        say(TRP_FAIL_TPL.format(name=type(e).__name__, detail=e))
        return
    types: dict = {}
    for k in ids.types:
        types[k] = TRP_TYPES[k]
    paths.write_json(paths.WriteJsonIn(path=OUT_TR_PROV, payload={
        K_SOURCE: TRP_SRC_URL, K_FETCHED: today_iso(),
        K_TYPES: types,
        K_BY_PROV: by_prov, K_LATEST_REF_PER: latest,
        K_NOTE: TRP_NOTE,
    }, indent=INDENT_1))
    on = by_prov.get(PROV_ON, {}).get(latest, {})
    say(TRP_DONE_TPL.format(n=len(by_prov), q=TRP_QUARTERS, out=OUT_TR_PROV.name))
    say(TRP_ROW_TPL.format(ref=latest, study=on.get(K_STUDY_ONLY, 0),
                           work=on.get(K_WORK_ONLY, 0), both=on.get(K_WORK_STUDY, 0)))


# =========================================================================
# 5. 省移民难度因子重算(清洗横切层 04e 的包装;本域只负责按序跑起来)
# =========================================================================


def build_clean_difficulty() -> None:
    """省移民难度指数重算 —— 调 etl/clean/04e_difficulty.py 子进程。

    04e **不溶进本域**:clean/ 是清洗横切层(一个关注点一个脚本、跨源生效),它消费本域
    前三步的 raw(statcan_tr_prov)+ 人工配额表 + pnp 域 draws,产出 processed/difficulty.json;
    11_build_stats 读它挂进 mart。本域只负责按序把它跑起来。
    子进程 stdout 直通(不 capture),与旧 _steps 一模一样;返回非零 → 抛错,
    「一步失败中止本轮」的语义与旧 _steps 逐字相同。
    """
    proc = subprocess.run([sys.executable, str(DIFFICULTY_SCRIPT)])
    if proc.returncode != 0:
        raise RuntimeError(DIFFICULTY_FAIL_TPL.format(code=proc.returncode))


# =========================================================================
# 6. PGWP 规则库(B1-4;quote-anchored,引用消失即保留旧表 exit 1)
# =========================================================================


def norm_pgwp(t: str) -> str:
    """归一化后再比对:去 md 强调星号、弯引号→直引号、压空白 —— 引用核对不被排版噪音干扰。"""
    t = t.replace(PGWP_STAR, "").replace(QUOTE_CURLY_RIGHT, QUOTE_STRAIGHT)
    t = t.replace(QUOTE_CURLY_LEFT, QUOTE_STRAIGHT)
    return WS_RE.sub(SPACE, t)


def fetch_norm(url: str) -> str:
    """实抓一页 → md → 归一化全文(引用逐字核验的比对底本)。"""
    html = httpx.get(url, headers={HDR_UA: PGWP_UA}, follow_redirects=True,
                     timeout=PGWP_TIMEOUT_S).text
    md = convert_md(ConvertIn(html=html, url=url, selector=None, removes=()))
    return norm_pgwp(md)


def to_pgwp_req(x: PgwpReqIn) -> dict:
    """一条人抄的规则 → 落盘门槛行(形状对齐 raw/pnp/<省>-req.json)。"""
    return {
        K_STREAM: x.rule[K_STREAM], K_SUBJECT: SUBJECT_APPLICANT, K_FACTOR: x.rule[K_FACTOR],
        K_OP: x.rule[K_OP], K_VALUE: x.rule[K_VALUE], K_VALUE_TEXT: x.rule[K_QUOTE],
        K_UNIT: x.rule[K_UNIT], K_BASIS: x.rule[K_BASIS], K_LABEL: x.rule[K_LABEL],
        K_SECTION: x.rule[K_PAGE], K_EFFECTIVE: x.rule.get(K_EFFECTIVE, ""), K_URL: x.url,
    }


def pgwp_missing_lines(missing: list) -> list:
    """引用核验未过的逐条明细(quote 截前 80 字)。"""
    lines = []
    for r in missing:
        lines.append(PGWP_MISSING_ROW_TPL.format(factor=r[K_FACTOR], stream=r[K_STREAM],
                                                 quote=r[K_QUOTE][:PGWP_QUOTE_CLIP]))
    return lines


def build_ircc_pgwp_rules() -> None:
    """联邦 PGWP 规则库(B1-4,2026-08-03;设计 docs/design/PGWP规则库-20260803.md)。

    **quote-anchored**(Frank 拍板「原文为准」):规则表每行带官方原文引用(valueText),
    本步每轮实抓 about/eligibility 两页,**逐条验证引用仍逐字存在于页面**——
    页面改版引用消失 → 保留旧表 + exit 1(钉 ircc 役末尾,红了触发 healthchecks 报警;
    crawl 役的 fed-pgwp 地图 diff 是第二道雷达)。规则是人工从原文抄的,机器管的是「原文没变」。
    """
    say(PGWP_PRINT_OUT_TPL.format(path=OUT_PGWP))
    pages = {PGWP_PAGE_ABOUT: fetch_norm(PGWP_URL_ABOUT),
             PGWP_PAGE_ELIG: fetch_norm(PGWP_URL_ELIG)}
    urls = {PGWP_PAGE_ABOUT: PGWP_URL_ABOUT, PGWP_PAGE_ELIG: PGWP_URL_ELIG}
    missing = []
    for r in PGWP_RULES:
        if norm_pgwp(r[K_QUOTE]) not in pages[r[K_PAGE]]:
            missing.append(r)
    if len(missing) > 0:
        fail_keep_old(FailIn(header=PGWP_MISSING_TPL.format(n=len(missing), total=len(PGWP_RULES)),
                             lines=pgwp_missing_lines(missing)))
    reqs = []
    for r in PGWP_RULES:
        reqs.append(to_pgwp_req(PgwpReqIn(rule=r, url=urls[r[K_PAGE]])))
    paths.write_json(paths.WriteJsonIn(path=OUT_PGWP, payload={
        K_PROVINCE: PROVINCE_FED, K_PROGRAM: PGWP_PROGRAM, K_URL: PGWP_URL_ABOUT,
        K_FETCHED: today_iso(),
        K_NOTE: PGWP_NOTE,
        K_REQUIREMENTS: reqs,
    }, indent=INDENT_1))
    say(PGWP_DONE_TPL.format(n=len(reqs), out=OUT_PGWP.name))


# =========================================================================
# 7. 联邦段官方规费(G8 v1;段落定位 + 交叉自校硬闸)
# =========================================================================


def fees_text(html: str) -> str:
    """费用页 HTML → 压平的正文(先拆掉导航/脚本等噪音容器)。"""
    soup = BeautifulSoup(html, PARSER_HTML)
    for t in soup(FEES_DROP_TAGS):
        t.decompose()
    return WS_RE.sub(SPACE, soup.get_text(SPACE, strip=True))


def fee_row(x: FeeRowIn) -> dict:
    """一条费用 → 落盘门槛行(键序即文件契约;费用与 TEER/NOC/地域无关,恒空)。"""
    return {K_STREAM: x.stream, K_SUBJECT: SUBJECT_APPLICANT, K_FACTOR: FEE_FACTOR, K_OP: FEE_OP,
            K_VALUE: x.amount, K_VALUE_TEXT: x.value_text, K_UNIT: FEE_UNIT,
            K_APPLIES_TEER: [], K_APPLIES_NOC: "", K_EXCLUDES_NOC: "", K_APPLIES_AREA: "",
            K_FAMILY_SIZE: None, K_BASIS: "", K_LABEL: x.label,
            K_SECTION: x.section, K_URL: FEES_URL}


def section_items(x: SectionItemsIn) -> ItemsOut:
    """Economic immigration 一节的四个条目(任一没解析到即记问题)。"""
    reqs = []
    problems = []
    for pat, stream, label in FEES_ITEMS:
        m = pat.search(x.seg)
        if m is None:
            problems.append(FEES_PROBLEM_ITEM_TPL.format(stream=stream))
            continue
        amount = int(m.group(1).replace(COMMA, ""))
        reqs.append(fee_row(FeeRowIn(stream=stream, amount=amount,
                                     value_text=m.group(0)[:FEES_VALUE_CLIP], label=label,
                                     section=FEES_SECTION)))
    return ItemsOut(reqs=reqs, problems=problems)


def first_group_of(m: object) -> str:
    """第一个非空捕获组(原 next(genexp) 的显式化)。"""
    for g in m.groups():
        if g:
            return g
    return ""


def bio_items(txt: str) -> ItemsOut:
    """生物识别两档(全文范围搜,不限在 Economic 一节里)。"""
    reqs = []
    problems = []
    for regex, stream, label in FEES_BIO_ITEMS:
        m = regex.search(txt)
        if m is None:
            problems.append(FEES_PROBLEM_ITEM_TPL.format(stream=stream))
            continue
        amount = int(first_group_of(m).replace(COMMA, ""))
        reqs.append(fee_row(FeeRowIn(stream=stream, amount=amount,
                                     value_text=m.group(0)[:FEES_VALUE_CLIP], label=label,
                                     section=FEES_SECTION_BIO)))
    return ItemsOut(reqs=reqs, problems=problems)


def fees_by_stream(reqs: list) -> dict:
    """条目名 → 金额(交叉自校与收尾报数共用)。"""
    by: dict = {}
    for r in reqs:
        by[r[K_STREAM]] = r[K_VALUE]
    return by


def rprf_problems(reqs: list) -> list:
    """交叉自校:principal - principalNoRprf = 永居权费,应为正数且 ≤ principal 的一半
    (改版最容易先烂在这)。"""
    by = fees_by_stream(reqs)
    problems = []
    if STREAM_PRINCIPAL in by and STREAM_PRINCIPAL_NO_RPRF in by:
        rprf = by[STREAM_PRINCIPAL] - by[STREAM_PRINCIPAL_NO_RPRF]
        if not (0 < rprf < by[STREAM_PRINCIPAL_NO_RPRF]):
            problems.append(FEES_PROBLEM_RPRF_TPL.format(principal=by[STREAM_PRINCIPAL],
                                                         no_rprf=by[STREAM_PRINCIPAL_NO_RPRF],
                                                         rprf=rprf))
    return problems


def fees_problem_lines(problems: list) -> list:
    """自校未过的逐条明细。"""
    lines = []
    for p in problems:
        lines.append(FEES_BULLET_TPL.format(problem=p))
    return lines


def build_ircc_fees() -> None:
    """联邦段官方规费(G8 v1,2026-08-03;案例库 C14「中介开价 3 万值吗」的拆账原料)。

    源 = IRCC 官方费用总表(ircc.canada.ca/english/information/fees/fees.asp,httpx 直连 200)。
    只收 **Economic immigration (including Express Entry)** 一节(官方明示适用于 PNP/EE/AIP/RCIP)
    + 生物识别两档。**段落定位后逐项正则**,任何一项没解析到 → 保留旧表 exit 1(硬闸,照 build_pgwp)。
    """
    say(FEES_PRINT_OUT_TPL.format(path=OUT_FEES))
    html = httpx.get(FEES_URL, headers={HDR_UA: FEES_UA}, follow_redirects=True,
                     timeout=FEES_TIMEOUT_S).text
    txt = fees_text(html)
    i = txt.find(FEES_SECTION)
    problems: list = []
    reqs: list = []
    if i < 0:
        problems.append(FEES_PROBLEM_NO_SECTION_TPL.format(section=FEES_SECTION))
    else:
        got = section_items(SectionItemsIn(seg=txt[i:i + FEES_SEG_LEN]))
        reqs += got.reqs
        problems += got.problems
    bio = bio_items(txt)
    reqs += bio.reqs
    problems += bio.problems
    problems += rprf_problems(reqs)
    if len(problems) > 0:
        fail_keep_old(FailIn(header=FEES_FAIL_HEADER, lines=fees_problem_lines(problems)))
    paths.write_json(paths.WriteJsonIn(path=OUT_FEES, payload={
        K_PROVINCE: PROVINCE_FED, K_PROGRAM: FEES_PROGRAM,
        K_SOURCE: FEES_SOURCE,
        K_URL: FEES_URL, K_FETCHED: today_iso(),
        K_NOTE: FEES_NOTE,
        K_REQUIREMENTS: reqs,
    }, indent=INDENT_1))
    say(FEES_DONE_TPL.format(n=len(reqs), by=fees_by_stream(reqs)))
